import openpyxl

wb = openpyxl.load_workbook(
    "/sessions/upbeat-kind-allen/mnt/uploads/DAILY_PLANNER_2025.xlsx",
    data_only=False,
)
log = wb["Log"]
print("Header:", [log.cell(1, c).value for c in range(1, 7)])
print("Total rows in Log:", log.max_row)
print()
print("First 5 rows (raw value + number_format + data_type):")
for r in range(2, 7):
    a = log.cell(r, 1)
    print(
        f"  row {r}: value={repr(a.value):40s} "
        f"fmt={repr(a.number_format):25s} dtype={repr(a.data_type)}"
    )

print()
print("Spot samples across the file (and their displayed equivalent):")
import datetime as dt

def display(value, fmt):
    if isinstance(value, dt.datetime):
        # Excel codes: H = hour-no-leading-zero, HH = padded.
        # Show what Sheets would render given the format.
        if "H" in fmt and "HH" not in fmt:
            return value.strftime("%Y-%m-%d ") + str(value.hour) + value.strftime(":%M")
        return value.strftime("%Y-%m-%d %H:%M")
    return str(value)

samples = [2, 50, 200, 500, 1000, 2000, max(2, log.max_row - 5), log.max_row]
for r in samples:
    if r > log.max_row:
        continue
    a = log.cell(r, 1)
    c = log.cell(r, 3)
    rendered = display(a.value, a.number_format)
    print(
        f"  row {r}: raw={repr(a.value):30s} fmt={repr(a.number_format):25s} "
        f"rendered={rendered!r:25s} cat={repr(c.value)}"
    )

print()
# Find rows where the hour is 0-9 to see how they're stored.
print("Rows for hours 0-9 (first 10):")
shown = 0
for r in range(2, log.max_row + 1):
    a = log.cell(r, 1)
    if isinstance(a.value, dt.datetime) and a.value.hour < 10:
        rendered = display(a.value, a.number_format)
        print(
            f"  row {r}: raw={repr(a.value):30s} fmt={repr(a.number_format):25s} "
            f"rendered={rendered!r}"
        )
        shown += 1
        if shown >= 10:
            break

# Count all rows by year-month-day-hour to find duplicates in the BACKUP.
print()
print("Duplicate timestamps in the backup (top 20 by count):")
from collections import Counter

c = Counter()
for r in range(2, log.max_row + 1):
    a = log.cell(r, 1)
    if isinstance(a.value, dt.datetime):
        key = a.value.replace(minute=0, second=0, microsecond=0)
        c[key] += 1

dupes = [(k, v) for k, v in c.items() if v > 1]
dupes.sort(key=lambda x: -x[1])
print(f"  total unique hours: {len(c)}, total rows: {sum(c.values())}, duplicate hours: {len(dupes)}")
for k, v in dupes[:20]:
    print(f"  {k} -> {v} copies")
